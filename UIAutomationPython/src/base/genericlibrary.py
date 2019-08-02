'''
Created on 24-May-2019

@author: Veer

'''

from builtins import len
from datetime import datetime
import logging
from time import strftime
import time
from traceback import print_stack

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, \
    ElementNotVisibleException, ElementNotSelectableException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

import utilities.custom_logger as cl
import os
import openpyxl 
from openpyxl.drawing.text import Font
from openpyxl.styles.colors import RED
from dateutil.utils import today


class GenericLibrary():
    
    log= cl.customlogger(logging.DEBUG)
    
    def __init__(self, driver):
        self.driver=driver
        
    def getByType(self,locatorType):
        locatorType=locatorType.lower()
        if locatorType=="id":
            return By.ID
        elif locatorType=="name":
            return By.NAME
        elif locatorType=="xpath":
            return By.XPATH
        elif locatorType=="css":
            return By.CSS_SELECTOR
        elif locatorType=="class":
            return By.CLASS_NAME
        elif locatorType=="link":
            return By.LINK_TEXT
        else:
            self.log.info("Locator type "+locatorType +" not correct/supported")
        return False

    def launchBrowser(self,browser):
        
        try:
            browser=browser.lower()
            if browser=="chrome":
                self.driver = webdriver.Chrome()
                self.log.info("****======================================== Execution Start ========================================****")
                self.log.info("chrome launched successfully")
                
                return self.driver;
                
            elif browser=="firefox":
                self.driver = webdriver.Firefox()
                self.log.info("****======================================== Execution Start ========================================****")
                self.log.info("Firefox launched successfully")
                return self.driver
                
            
            elif browser=="ie":
                self.driver = webdriver.Ie()
                self.log.info("****======================================== Execution Start ========================================****")
                self.log.info("Internet explorer launched successfully")
                return self.driver
                
        except:
            print("browser not found")
            print_stack()
        
        
    def sleepTime(self, sec):    
            time.sleep(sec)
    
    def quitBrowser(self):    
            self.driver.quit()
            

    def closeBrowser(self): 
            time.sleep(10)   
            self.driver.close()
            self.log.info("browser close successfully")
            self.log.info("****======================================== Execution End ==========================================****")
            
    def explicitlyWait(self, locator, timeout):    
        #self.driver.implicitly_wait(second)
            
        element = None
        try:
#             byType = self.getByType(locatorType)
            self.driver.getElement(locator)
            print("waiting for maximum ::"+str(timeout)+" :: seconds for elements to be clickable")
            wait = WebDriverWait(self.driver,10,poll_frequency=1, ignored_exceptions=[NoSuchElementException, ElementNotVisibleException, ElementNotSelectableException])
            
            #element=wait.until(EC.element_attribute_should_be(byType,"stopFilter_stops-0"))
            element=wait.until().element_should_be_visible(locator)
            print("Element appeared on the web page")
            
        except:
            print("element not appeared on the web page")
            print_stack()
        return element
    
            
    def maxWindow(self):    
            self.driver.maximize_window()
        
    def getUrl(self, baseuri):
        try:
            self.driver.get(baseuri)
            self.log.info("navigate to Web URL :"+baseuri)
        except:
            self.log.info("Base URL does not found")
            print_stack()
        
    
    def getElement(self, locator, locatorType="id"):
        
        element = None
        try:
            locatorType= locatorType.lower()
            byType=self.getByType(locatorType)
            element = self.driver.find_element(byType,locator)
            self.log.info("Element found with locator : "+ locator +" and locator type : "+locatorType)
        except:
            self.log.info("Element not found with this locator : "+ locator +" and locator type : "+locatorType)
        return element
    
    def elementClick(self,locator, locatorType="id"):
        try:
            element= self.getElement(locator,locatorType)
            element.click()
            self.log.info("clicked on element with locator : "+ locator+ " locator type : "+ locatorType)
        except:
            self.log.info("cannot click on the element with locator : "+locator + "locator type : "+ locatorType)
            print_stack()
            
            
    def sendkeys(self, data, locator, locatorType="id"):
        try:
            element=self.getElement(locator,locatorType)
            element.send_keys(data)
            self.log.info("send data on the element with locator :"+locator + " locator type :"+ locatorType)
        except:
            self.log.info("can not send data on the element with locator :"+locator + " locator type :"+ locatorType)
            print_stack()
            
            
    def isElementPresent(self, locator, locatorType="id"):
        try:
            element= self.driver.getElement(locator, locatorType)
            if element is not None:
                print("Element found")
                return True
            else:
                print("Element Not found")
                return False
        except:
            print("Element not found")
            return False
        
        ''' here is method of uodate status in excel '''
        
    def writedataexcel(self, rownum, cellvalue):
        
        try:
                
            wb = openpyxl.load_workbook('C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/TestData/RegressionScenarios.xlsx')
              
            sheet = wb.active
            sheet.title = "Sheet1" 
            # sheet.cell(row = 1, column = 9).font = Font(Color = RED)
            sheet.cell(row = rownum, column = 8).value = cellvalue
              
            wb.save('C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/TestData/RegressionScenarios.xlsx') 
            print("updated sheet")

        except:
            print("record is not updated")
            print_stack()
        
        
    def elementPresentCheck(self, locator, byType):
        try:
            elementList =self.driver.find_element(byType, locator)
            if len(elementList)>0:
                print("Element found")
                return True
            else:
                print("Element not found")
                return False
        except:
            print("Element not found")
            return False
        
        
    def takeScreenshot(self, messages):
        somenum = str(round(time.time()*100))
#         currentdate = os.makedirs(datetime.now().strftime("%H_%M_%S"))
        
        currenttime = messages + "."+ datetime.now().strftime("_%H_%M_%S_")+somenum
        #Filename = str(round(time.time()*1000))+".png"
        Filename =  str(currenttime)+".png"
        screenshotDirectory = ("C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/Screenshots/"+Filename)
        destinationFile = screenshotDirectory
    
        try:
            self.driver.save_screenshot(destinationFile)
            self.log.info("Screenshot save to directory --> ::"+destinationFile)
        except :
            self.log.error("### exception occurred when taking screen shot")
            print_stack()
            print("not a directory issue")
            
            
    def screenShots(self, resultMessage):
        fileName = resultMessage + "." + str(round(time.time()*100))+".png"
#         currenttime = datetime.now().strftime("%H_%M_%S")+fileName

        screenshotDirectory = "../screenShots/"
        relativeFileName = screenshotDirectory+ fileName
        currentDirectory = os.path.dirname(__file__)
        destinationFile = os.path.join(currentDirectory,relativeFileName)
        destinationDirectory = os.path.join(currentDirectory,screenshotDirectory)
        
        try:
            if not os.path.exists(destinationDirectory):
                os.makedirs(destinationDirectory)
                
                self.driver.save_screenshot(destinationFile)
                self.log.info("Screenshot save to directory : "+destinationFile)
        except:
            self.log.error("### exception occurred when taking screen shot")
            print_stack()
        
        
            
            
    def waitForElement(self, locator, locatorType="id", timeout=10, pollFrequency=0.5):
        element = None
        try:
            byType = self.getByType(locatorType)
            print("waiting for maximum ::"+str(timeout)+" :: seconds for elements to be clickable")
            wait = WebDriverWait(self.driver,10,poll_frequency=1, ignored_exceptions=[NoSuchElementException, ElementNotVisibleException, ElementNotSelectableException])
            
#             element=wait.until(EC.element_attribute_should_be(byType,"stopFilter_stops-0"))
            print("Element appeared on the web page")
            
        except:
            print("element not appeared on the web page")
            print_stack()
        return element
            
            
            
    def webScroll(self, direction="up"):
        if direction=="up":
            self.driver.execute_script("window.scrollBy(0,-1000);")
            
        if direction=="down":
            self.driver.execute_script("window.scrollBy(0,1000);")
        
            
            
            
            
            
            
            
            
            
            
            
            
            
            
        