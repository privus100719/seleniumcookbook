'''
Created on 26-May-2019

@author: Veer
'''

import openpyxl 

wb = openpyxl.load_workbook('C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/TestData/RegressionScenarios.xlsx')

## Read the sheets name for your entire workbook
wb.get_sheet_names()

## Create a variable for each sheet in the work book, to manipulate each sheet
sheet1 = wb.get_sheet_by_name('Sheet1')

for row in range(1, sheet1.max_row+1):
    cell = sheet1.cell(row=row, column=1)
    if cell.value is not None:
        cell.value = 'hi sheeteee' + cell.value

wb.save('C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/TestData/RegressionScenarios.xlsx')

print("updated sheet")