'''
Created on 26-May-2019

@author: Veer
'''
# import openpyxl module 
import openpyxl 
from openpyxl.drawing.text import Font
from openpyxl.styles.colors import RED
# from openpyxl.drawing.text.Color import Color
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 

class WriteData():
    
    def writedataexcel(self, rownum, cellvalue):
        

        wb = openpyxl.load_workbook('C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/TestData/RegressionScenarios.xlsx')
          
        sheet = wb.active
        sheet.title = "Sheet1"
          
        # writing to the specified cell 
        # sheet.cell(row = 1, column = 9).font = Font(Color = RED)
        sheet.cell(row = rownum, column = 8).value = cellvalue
          
        # sheet.cell(row = 2, column = 2).value = ' everyone two jkh'
          
        # set the height of the row 
        # sheet.row_dimensions[1].height = 70
        #   
        # # set the width of the column 
        # sheet.column_dimensions['B'].width = 20
          
        # save the file 
        wb.save('C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/TestData/RegressionScenarios.xlsx') 
        print("updated sheet")



w1 = WriteData()

w1.writedataexcel(2, 'PASS')

# if __name__ == '__main__':
#     main()