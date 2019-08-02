'''
Created on 26-May-2019

@author: Veer
'''

import openpyxl
from xlrd.timemachine import xrange

def main():
    book = openpyxl.load_workbook('C:/Users/rajve/workspace/Oxygen Workspace/MyBSFUIAutomationPython/TestData/RegressionScenarios.xlsx')
    print (book.get_sheet_names())
    # ['Sheet2', 'New Title', 'Sheet1']
    # Get a sheet to read
    sheet = book.get_sheet_by_name('Sheet1')
    # No of written Rows in sheet
    r = sheet.max_row
    # No of written Columns in sheet
    c = sheet.max_column
    # Reading each cell in excel
    for i in xrange(1, r+1):
        for j in xrange(1, c+1):
            print(sheet.cell(row=i, column=j).value)

if __name__ == '__main__':
    main()